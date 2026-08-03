"""Build SmartRecruiters or Workable local review queues from workbook and VPS records."""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook


PLATFORMS = {"smartrecruiters", "workable"}


def canonical_url(value: object, platform: str) -> str | None:
    parsed = urlsplit(str(value or "").strip())
    parts = [part for part in parsed.path.split("/") if part]
    if platform == "smartrecruiters":
        if parsed.hostname not in {"jobs.smartrecruiters.com", "www.smartrecruiters.com"}:
            return None
        if len(parts) < 2:
            return None
        path = "/" + "/".join(parts[:2])
        return urlunsplit(("https", "jobs.smartrecruiters.com", path, "", ""))
    if platform == "workable":
        if parsed.hostname != "apply.workable.com":
            return None
        if len(parts) >= 3 and parts[1].lower() in {"j", "jobs"}:
            path = "/" + "/".join(parts[:3])
        elif len(parts) >= 2 and parts[0].lower() in {"j", "jobs"}:
            path = "/" + "/".join(parts[:2])
        else:
            return None
        return urlunsplit(("https", "apply.workable.com", path, "", ""))
    raise ValueError(f"unsupported platform: {platform}")


def build_records(workbook_path: Path, platform: str) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows = workbook.active.iter_rows(values_only=True)
    headers = tuple(str(value or "").strip() for value in next(rows))
    required = {"Company", "Job Title", "Location", "Application URL"}
    missing = required.difference(headers)
    if missing:
        raise ValueError(f"workbook is missing required headers: {sorted(missing)}")
    indexes = {header: headers.index(header) for header in required}
    imported_at = datetime.now(timezone.utc).isoformat()
    records: list[dict[str, object]] = []
    seen: set[str] = set()
    for row in rows:
        url = canonical_url(row[indexes["Application URL"]], platform)
        if url is None or url in seen:
            continue
        seen.add(url)
        records.append(
            {
                "company": str(row[indexes["Company"]] or "").strip(),
                "role": str(row[indexes["Job Title"]] or "").strip(),
                "job_url": url,
                "status": "NOT_ATTEMPTED",
                "failure_reason": f"Imported from {platform} product jobs workbook",
                "missing_required": "",
                "updated_at": imported_at,
                "location": str(row[indexes["Location"]] or "").strip(),
            }
        )
    return records


def merge_records(base: list[dict[str, object]], incoming: list[dict[str, object]], platform: str):
    merged = {str(record["job_url"]): dict(record) for record in base}
    for record in incoming:
        url = canonical_url(record.get("job_url"), platform)
        if url is None:
            continue
        updated = dict(merged.get(url, {}))
        updated.update({str(key): value for key, value in record.items() if value not in (None, "", [])})
        updated["job_url"] = url
        merged[url] = updated
    return list(merged.values())


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--platform", choices=sorted(PLATFORMS), required=True)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--merge-json", action="append", default=[], type=Path)
    args = parser.parse_args()
    records = build_records(args.workbook, args.platform)
    for path in args.merge_json:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(payload, list):
            raise ValueError(f"merge JSON must contain a list: {path}")
        records = merge_records(records, [item for item in payload if isinstance(item, dict)], args.platform)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(records, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"platform={args.platform} records={len(records)} output={args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

"""Build an Ashby failed-JSON worker source from an Excel job tracker."""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime, timezone
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook


REQUIRED_HEADERS = {
    "Posting Date",
    "Company",
    "Job Title",
    "Location",
    "Application URL",
}


def canonical_ashby_url(value: object) -> str | None:
    """Return a query-free Ashby job URL, or None for invalid input."""
    raw = str(value or "").strip()
    if not raw:
        return None
    parsed = urlsplit(raw)
    if parsed.scheme.lower() not in {"http", "https"}:
        return None
    if parsed.hostname not in {"jobs.ashbyhq.com", "ashbyhq.com"}:
        return None
    path = parsed.path.rstrip("/")
    if len([part for part in path.split("/") if part]) < 2:
        return None
    return urlunsplit(("https", "jobs.ashbyhq.com", path, "", ""))


def _timestamp(value: object) -> str:
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, datetime.min.time())
    else:
        return datetime.now(timezone.utc).isoformat()
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc).isoformat()


def build_records(workbook_path: Path) -> list[dict[str, str]]:
    """Read all unique valid Ashby jobs while preserving workbook order."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = tuple(str(value or "").strip() for value in next(rows))
    missing = REQUIRED_HEADERS.difference(headers)
    if missing:
        raise ValueError(f"workbook is missing required headers: {sorted(missing)}")
    indexes = {header: headers.index(header) for header in REQUIRED_HEADERS}

    records: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in rows:
        url = canonical_ashby_url(row[indexes["Application URL"]])
        if url is None or url in seen:
            continue
        seen.add(url)
        records.append(
            {
                "company": str(row[indexes["Company"]] or "").strip(),
                "role": str(row[indexes["Job Title"]] or "").strip(),
                "job_url": url,
                "status": "NOT_ATTEMPTED",
                "failure_reason": "Imported from Ashby product management workbook",
                "missing_required": "",
                "updated_at": _timestamp(row[indexes["Posting Date"]]),
                "location": str(row[indexes["Location"]] or "").strip(),
            }
        )
    return records


def merge_records(
    workbook_records: list[dict[str, str]],
    additional_records: list[dict[str, object]],
) -> list[dict[str, object]]:
    """Merge actual attempt records over workbook placeholders by canonical URL."""
    merged: dict[str, dict[str, object]] = {
        record["job_url"]: dict(record) for record in workbook_records
    }
    for incoming in additional_records:
        url = canonical_ashby_url(incoming.get("job_url"))
        if url is None:
            continue
        current = merged.get(url, {})
        updated = dict(current)
        for key, value in incoming.items():
            if value not in (None, "", []):
                updated[str(key)] = value
        updated["job_url"] = url
        merged[url] = updated
    return list(merged.values())


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--merge-json", action="append", default=[], type=Path)
    args = parser.parse_args()
    records: list[dict[str, object]] = list(build_records(args.workbook))
    for merge_path in args.merge_json:
        payload = json.loads(merge_path.read_text(encoding="utf-8"))
        if not isinstance(payload, list):
            raise ValueError(f"merge JSON must contain a list: {merge_path}")
        records = merge_records(records, [item for item in payload if isinstance(item, dict)])
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(records, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    print(f"records={len(records)} output={args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

"""Build a Lever failed/unsubmitted review queue from workbook and VPS records."""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime, timezone
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook


HEADERS = ("Posting Date", "Company", "Job Title", "Location", "Application URL")
LEVER_HOSTS = {"jobs.lever.co", "jobs.eu.lever.co"}


def canonical_lever_url(value: object) -> str | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    parsed = urlsplit(raw)
    if parsed.scheme.lower() not in {"http", "https"} or parsed.hostname not in LEVER_HOSTS:
        return None
    parts = [part for part in parsed.path.split("/") if part]
    if parts and parts[-1].lower() == "apply":
        parts.pop()
    if len(parts) < 2:
        return None
    return urlunsplit(("https", parsed.hostname, "/" + "/".join(parts[:2]), "", ""))


def _timestamp(value: object) -> str:
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, datetime.min.time())
    else:
        return datetime.now(timezone.utc).isoformat()
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc).isoformat()


def build_records(workbook_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows = workbook.active.iter_rows(values_only=True)
    headers = tuple(str(value or "").strip() for value in next(rows))
    missing = set(HEADERS).difference(headers)
    if missing:
        raise ValueError(f"workbook is missing required headers: {sorted(missing)}")
    indexes = {header: headers.index(header) for header in HEADERS}
    records: list[dict[str, object]] = []
    seen: set[str] = set()
    for row in rows:
        url = canonical_lever_url(row[indexes["Application URL"]])
        if url is None or url in seen:
            continue
        seen.add(url)
        records.append(
            {
                "company": str(row[indexes["Company"]] or "").strip(),
                "role": str(row[indexes["Job Title"]] or "").strip(),
                "job_url": url,
                "status": "NOT_ATTEMPTED",
                "failure_reason": "Imported from Lever product management workbook",
                "missing_required": "",
                "updated_at": _timestamp(row[indexes["Posting Date"]]),
                "location": str(row[indexes["Location"]] or "").strip(),
            }
        )
    return records


def merge_records(
    base: list[dict[str, object]], incoming_records: list[dict[str, object]]
) -> list[dict[str, object]]:
    merged = {str(record["job_url"]): dict(record) for record in base}
    for incoming in incoming_records:
        url = canonical_lever_url(incoming.get("job_url"))
        if url is None:
            continue
        updated = dict(merged.get(url, {}))
        updated.update(
            {str(key): value for key, value in incoming.items() if value not in (None, "", [])}
        )
        updated["job_url"] = url
        merged[url] = updated
    return list(merged.values())


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--merge-json", action="append", default=[], type=Path)
    args = parser.parse_args()
    records = build_records(args.workbook)
    for path in args.merge_json:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(payload, list):
            raise ValueError(f"merge JSON must contain a list: {path}")
        records = merge_records(records, [item for item in payload if isinstance(item, dict)])
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(records, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"records={len(records)} output={args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
